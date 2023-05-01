from openpyxl import load_workbook
import numpy as np
import matplotlib as mpl
import matplotlib.pyplot as plt
from scipy.optimize import leastsq
from scipy.optimize import curve_fit

wb = load_workbook(filename="/home/wqy/Downloads/s-t.xlsx")


sheets = wb.sheetnames

sheet_first = sheets[0]

ws = wb[sheet_first]

rows = ws.rows
columns = ws.columns

#for col in columns:
#    data_col = [ row.value for row in col ]
#    print(data_col)

#data_25_col = [ ws.cell(row=3 + x, column=1).value for x in range(1,100)]


def col_to_list(start_row, start_col, end_col):
    col_list = []
    for col in ws.iter_cols(min_row=start_row, min_col=start_col, max_col=end_col):
        for cell in col:
            if cell.value is not None:
                col_list.append(cell.value)
    return col_list


def func(x, a, b):
    return a*np.exp(b/x)


t_25 = col_to_list(4, 1, 1)
s_25 = col_to_list(4, 2, 2)

np_t_25 = np.array(t_25)
np_s_25 = np.array(s_25)

p = np.polynomial.Polynomial.fit(np_t_25, np_s_25, deg=10)
plt.plot(np_t_25, np_s_25, 's', label='original values')
plt.plot(np_t_25, p(np_t_25), 'r', label='Power series')

plt.legend()
plt.show()
print(p)

#popt, pcov = curve_fit(func, t_25, s_25)
#a = popt[0]
#b = popt[1]
#print("a:", a, "\n")
#print("b:", b, "\n")
#yvals = func(t_25, a, b)

#plot1 = plt.plot(t_25, s_25, 'x')
#plot2 = plt.plot(t_25, yvals, 'r', label='curve_fit values')
#plt.legend(loc=4)
#plt.show()

#t_40 = col_to_list(4, 3, 3)
#s_40 = col_to_list(4, 4, 4)

#t_50 = col_to_list(4, 5, 5)
#s_50 = col_to_list(4, 6, 6)

#t_60 = col_to_list(4, 7, 7)
#s_60 = col_to_list(4, 8, 8)


#print(s_60);
#print(ws.cell(row=3, column=1).value)
#print(data_25_col)
